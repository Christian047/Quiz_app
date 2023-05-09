# Create an excel file that will read and write, and also convert a text to an excel file.
import openpyxl
import csv

# we Create a new workbook object using the "workbook()" constructor
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write data to cells
worksheet['A1'] = 'hello'
worksheet['B1'] = 'world'

# Save the workbook
workbook.save('cohort3.xlsx')


# from line 5 to 18 we create a new empty Excel workbook called "cohort3.xlsx"
# which contains a singleworksheet with the text "hello" in cell A1
#  and "world" in cell B1.

# to read we
# Load the workbook
workbook = openpyxl.load_workbook('cohort3.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Read data from cells
cell_value1 = worksheet['A1'].value
cell_value2 = worksheet['B1'].value

# Print the data
print(cell_value1, cell_value2)


# This will load the "cohort3.xlsx" file we created earlier and read
#   the values from cells A1 and B1. The values are then printed to the console.



#  to write we
# Load the workbook
workbook = openpyxl.load_workbook('cohort3.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Write new data to cells
worksheet['C1'] = 'how'
worksheet['D1'] = 'are'
worksheet['E1'] = 'you?'

# Save the workbook
workbook.save('cohort3.xlsx')


# we load the "cohort3.xlsx" file we created earlier 
# and add new data to cells C1, D1, and E1. 
# The modified workbook is then saved.

#  Converting a text file to an Excel file. first we import csv which we did from the begining
# Load the text file
with open('file.txt', 'r') as infile:
    reader = csv.reader(infile, delimiter='\t')
    data = [row for row in reader]

# Create a new workbook
workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Write data to cells
for row in data:
    worksheet.append(row)

# Save the workbook
workbook.save('cohort.xlsx')


# This will load the text file "file.txt", which is assumed to be a tab-delimited
# file, and convert it into an Excel file "output.xlsx" where each row of the text
# file becomes a row in the Excel spreadsheet.

